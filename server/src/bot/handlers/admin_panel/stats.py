"""
Админ-панель для управления ботом
"""

from aiogram import Dispatcher, types, F
from aiogram.filters import Command, StateFilter
from aiogram.fsm.context import FSMContext
from aiogram.utils.keyboard import InlineKeyboardBuilder
from aiogram.types import InlineKeyboardButton
from loguru import logger

from datetime import datetime, timedelta
from io import BytesIO
from openpyxl import Workbook

from src.bot.database.states import AdminPanelState, PostAttachState
from src.bot.database.methods import (
    get_ad_by_id, update_ad, mark_ad_removed,
    set_moderator, get_user_by_id, is_moderator,
    count_users, count_banned, get_users_csv_rows,
    add_to_blacklist, remove_from_blacklist, is_banned, get_user_by_username, get_user_by_tg_id,
    get_details_stats_aggregated, get_contact_stats_aggregated,
    get_details_detailed_rows, get_contact_detailed_rows,
    get_total_placed_sold_removed, get_top_placed, get_top_sold, get_top_removed,
    get_top_reviews_activity,
    count_trusted_sellers, set_trusted_seller, is_trusted_seller,
)
from src.bot.settings.settings import ADMIN_IDS, CHANNEL_ID, CHANNEL_USERNAME, BOT_USERNAME
from src.bot.settings.constants import CONDITIONS, DEFAULT_CITIES
from src.bot.loader import bot
from src.bot.keyboards.keyboards import ad_in_channel_kb
from sqlalchemy import select, func
from src.bot.database.methods import async_session
from src.models import User, Ad, Review
from src.bot.middlewares.throttle_middleware import invalidate_banned_cache


# === /post_attach (пост в канал с кнопкой на бота) ===

from ._common import *

# === СТАТИСТИКА ===

def _parse_date_dmy(s: str):
    """Парсинг даты ДД.ММ.ГГГГ. Возвращает datetime или None."""
    s = (s or "").strip()
    try:
        return datetime.strptime(s, "%d.%m.%Y")
    except ValueError:
        return None


def _parse_period(s: str):
    """Парсинг интервала вида 30.01.2026-20.02.2026. Возвращает (date_start, date_end) или (None, None). Даты могут быть в любом порядке — проверка start <= end в обработчике."""
    s = (s or "").strip()
    if "-" not in s:
        return None, None
    a, b = s.split("-", 1)
    d1, d2 = _parse_date_dmy(a.strip()), _parse_date_dmy(b.strip())
    if not d1 or not d2:
        return None, None
    return d1, d2


async def admin_stats_menu(callback: types.CallbackQuery, state: FSMContext):
    """Меню «Статистика»: тип статистики."""
    await callback.answer()
    if not await check_admin_rights(callback.from_user.id):
        return
    await state.clear()
    text = "Выберите тип статистики:"
    keyboard = InlineKeyboardBuilder()
    keyboard.row(InlineKeyboardButton(text="📈 Статистика переходов", callback_data="admin:stats:transitions"))
    keyboard.row(InlineKeyboardButton(text="⭐ Рейтинги", callback_data="admin:stats:ratings"))
    keyboard.row(InlineKeyboardButton(text="⬅️ Отмена", callback_data="admin:back"))
    try:
        await callback.message.edit_text(text, reply_markup=keyboard.as_markup())
    except Exception:
        await callback.message.answer(text, reply_markup=keyboard.as_markup())


async def admin_stats_transitions_menu(callback: types.CallbackQuery, state: FSMContext):
    """Меню «Статистика переходов»: за весь период / выбрать период."""
    await callback.answer()
    if not await check_admin_rights(callback.from_user.id):
        return
    await state.clear()
    text = "Выберите метод выгрузки статистики"
    keyboard = InlineKeyboardBuilder()
    keyboard.row(InlineKeyboardButton(text="За весь период", callback_data="admin:stats:trans:all"))
    keyboard.row(InlineKeyboardButton(text="Текущие сутки", callback_data="admin:stats:trans:today"))
    keyboard.row(InlineKeyboardButton(text="Выбрать период", callback_data="admin:stats:trans:period"))
    keyboard.row(InlineKeyboardButton(text="⬅️ Назад", callback_data="admin:stats"))
    try:
        await callback.message.edit_text(text, reply_markup=keyboard.as_markup())
    except Exception:
        await callback.message.answer(text, reply_markup=keyboard.as_markup())
        try:
            await callback.message.delete()
        except Exception:
            pass


async def _build_transitions_excel_async(date_from=None, date_to=None) -> bytes:
    """Асинхронно собирает Excel файл со статистикой переходов с тремя листами.
    Возвращает bytes содержимого Excel файла."""
    details_agg = await get_details_stats_aggregated(date_from, date_to)
    contact_agg = await get_contact_stats_aggregated(date_from, date_to)
    details_det = await get_details_detailed_rows(date_from, date_to)
    contact_det = await get_contact_detailed_rows(date_from, date_to)
    
    wb = Workbook()
    
    # Удаляем дефолтный лист
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # Лист 1: Переходы (Подробнее) - агрегированная статистика (имя: @username или tg_id)
    ws1 = wb.create_sheet("Переходы")
    ws1.append(["№", "id_пользователя_tg", "имя_пользователя", "счетчик_подробнее"])
    for i, (tg_id, username, cnt) in enumerate(details_agg, 1):
        name_display = f"@{username}" if username else str(tg_id)
        ws1.append([i, tg_id, name_display, cnt])
    for col_letter, width in [("A", 6), ("B", 22), ("C", 28), ("D", 20)]:
        ws1.column_dimensions[col_letter].width = width

    # Лист 2: Просмотр профиля (Профиль продавца) - агрегированная статистика (имя: @username или tg_id)
    ws2 = wb.create_sheet("Просмотр профиля")
    ws2.append(["№", "id_пользователя_tg", "имя_пользователя", "счетчик_профиль_продавца"])
    for i, (tg_id, username, cnt) in enumerate(contact_agg, 1):
        name_display = f"@{username}" if username else str(tg_id)
        ws2.append([i, tg_id, name_display, cnt])
    for col_letter, width in [("A", 6), ("B", 22), ("C", 28), ("D", 24)]:
        ws2.column_dimensions[col_letter].width = width

    # Лист 3: Детальная статистика. Время в БД хранится в UTC; в выгрузку выводим МСК (UTC+3)
    TZ_OFFSET_HOURS = 3
    ws3 = wb.create_sheet("Детальная статистика")
    ws3.append(["Дата_время", "пользователь", "действие", "id_объявления", "продавец"])
    # Детальная статистика переходов "Подробнее"
    for r in details_det:
        # r = (created_at, user_tg_id, username, ad_id, seller_username, seller_tg_id)
        user_display = f"@{r[2]}" if r[2] else str(r[1])
        seller_display = f"@{r[4]}" if r[4] and r[4].strip() else str(r[5]) if r[5] else ""
        ts = (r[0] + timedelta(hours=TZ_OFFSET_HOURS)) if r[0] else None
        ws3.append([
            ts.strftime("%Y-%m-%d %H:%M:%S") if ts else "",
            user_display,
            "Подробнее",
            r[3],
            seller_display
        ])
    # Детальная статистика просмотра профиля
    for r in contact_det:
        # r = (created_at, buyer_tg_id, buyer_username, "Профиль и контакты", ad_id, seller_name)
        user_display = f"@{r[2]}" if r[2] else str(r[1])
        ts = (r[0] + timedelta(hours=TZ_OFFSET_HOURS)) if r[0] else None
        ws3.append([
            ts.strftime("%Y-%m-%d %H:%M:%S") if ts else "",
            user_display,
            r[3],
            r[4],
            r[5]
        ])
    
    for col_letter, width in [("A", 22), ("B", 28), ("C", 26), ("D", 15), ("E", 28)]:
        ws3.column_dimensions[col_letter].width = width

    # Сохраняем в BytesIO
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


async def admin_stats_transitions_all(callback: types.CallbackQuery, state: FSMContext):
    """Статистика переходов за весь период — отправка Excel файла с тремя листами."""
    await callback.answer()
    if not await check_admin_rights(callback.from_user.id):
        return
    await state.clear()
    excel_content = await _build_transitions_excel_async()
    from aiogram.types import BufferedInputFile
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    
    file = BufferedInputFile(excel_content, filename=f"stats_transitions_{timestamp}.xlsx")
    keyboard = InlineKeyboardBuilder()
    keyboard.row(InlineKeyboardButton(text="⬅️ Назад", callback_data="admin:stats:transitions"))
    await callback.message.answer_document(
        document=file,
        caption="Статистика за весь период:",
        reply_markup=keyboard.as_markup(),
    )
    try:
        await callback.message.delete()
    except Exception:
        pass


async def admin_stats_transitions_today(callback: types.CallbackQuery, state: FSMContext):
    """Статистика переходов за текущие сутки (МСК, UTC+3)."""
    await callback.answer()
    if not await check_admin_rights(callback.from_user.id):
        return
    await state.clear()
    TZ_OFFSET_HOURS = 3
    now_msk = datetime.utcnow() + timedelta(hours=TZ_OFFSET_HOURS)
    day_start = now_msk.replace(hour=0, minute=0, second=0, microsecond=0) - timedelta(hours=TZ_OFFSET_HOURS)
    day_end = now_msk.replace(hour=23, minute=59, second=59, microsecond=999999) - timedelta(hours=TZ_OFFSET_HOURS)
    excel_content = await _build_transitions_excel_async(day_start, day_end)
    from aiogram.types import BufferedInputFile
    date_str = now_msk.strftime('%Y%m%d')
    file = BufferedInputFile(excel_content, filename=f"stats_transitions_today_{date_str}.xlsx")
    keyboard = InlineKeyboardBuilder()
    keyboard.row(InlineKeyboardButton(text="⬅️ Назад", callback_data="admin:stats:transitions"))
    await callback.message.answer_document(
        document=file,
        caption=f"Статистика переходов за {now_msk.strftime('%d.%m.%Y')} (МСК):",
        reply_markup=keyboard.as_markup(),
    )
    try:
        await callback.message.delete()
    except Exception:
        pass


async def admin_stats_transitions_period_start(callback: types.CallbackQuery, state: FSMContext):
    """Запрос интервала для статистики переходов (30.01.2026-20.02.2026)."""
    await callback.answer()
    if not await check_admin_rights(callback.from_user.id):
        return
    await state.set_state(AdminPanelState.stats_transitions_period_input)
    await state.update_data(trans_period_ask_msg_id=callback.message.message_id, trans_period_has_error=False)
    text = "Введите временной интервал для вывода статистики, в формате: 30.01.2026-20.02.2026"
    keyboard = InlineKeyboardBuilder()
    keyboard.row(InlineKeyboardButton(text="⬅️ Назад", callback_data="admin:stats:transitions"))
    try:
        await callback.message.edit_text(text, reply_markup=keyboard.as_markup())
    except Exception:
        await callback.message.answer(text, reply_markup=keyboard.as_markup())


async def admin_stats_transitions_period_input(message: types.Message, state: FSMContext):
    """Обработка ввода интервала для статистики переходов."""
    d_start, d_end = _parse_period(message.text)
    back_kb = InlineKeyboardBuilder()
    back_kb.row(InlineKeyboardButton(text="⬅️ Назад", callback_data="admin:stats:transitions"))
    
    # Удаляем сообщение пользователя
    try:
        await message.delete()
    except Exception:
        pass
    
    data = await state.get_data()
    period_ask_msg_id = data.get("trans_period_ask_msg_id")
    has_error_msg = data.get("trans_period_has_error", False)
    
    # Валидация формата
    if not d_start or not d_end:
        # Если уже было сообщение об ошибке — только удалили сообщение юзера, новое не отправляем
        if has_error_msg:
            return
        await state.update_data(trans_period_has_error=True)
        err = "Неверный формат. Укажите интервал в формате ДД.ММ.ГГГГ-ДД.ММ.ГГГГ (например 30.01.2026-20.02.2026)."
        if period_ask_msg_id:
            try:
                await bot.edit_message_text(
                    chat_id=message.chat.id,
                    message_id=period_ask_msg_id,
                    text=err,
                    reply_markup=back_kb.as_markup(),
                )
            except Exception:
                await message.answer(err, reply_markup=back_kb.as_markup())
        else:
            await message.answer(err, reply_markup=back_kb.as_markup())
        return
    
    # Валидация: дата начала должна быть меньше даты конца, одинаковые даты не допускаются
    if d_start >= d_end:
        if has_error_msg:
            return
        await state.update_data(trans_period_has_error=True)
        err = "Дата начала периода должна быть раньше даты окончания. Укажите интервал от меньшей даты к большей (например 01.01.2025-06.02.2026). Одинаковые даты не допускаются."
        if period_ask_msg_id:
            try:
                await bot.edit_message_text(
                    chat_id=message.chat.id,
                    message_id=period_ask_msg_id,
                    text=err,
                    reply_markup=back_kb.as_markup(),
                )
            except Exception:
                await message.answer(err, reply_markup=back_kb.as_markup())
        else:
            await message.answer(err, reply_markup=back_kb.as_markup())
        return
    
    # Если валидация прошла успешно, сбрасываем флаг ошибки
    await state.update_data(trans_period_has_error=False)
    
    # Удаляем сообщение бота с запросом интервала
    data = await state.get_data()
    period_ask_msg_id = data.get("trans_period_ask_msg_id")
    if period_ask_msg_id:
        try:
            await bot.delete_message(chat_id=message.chat.id, message_id=period_ask_msg_id)
        except Exception:
            pass
    
    # Генерируем Excel файл и отправляем
    date_end = d_end.replace(hour=23, minute=59, second=59, microsecond=999999)
    excel_content = await _build_transitions_excel_async(d_start, date_end)
    from aiogram.types import BufferedInputFile
    date_str = f"{d_start.strftime('%Y%m%d')}_{d_end.strftime('%Y%m%d')}"
    period_text = f"{d_start.strftime('%d.%m.%Y')} - {d_end.strftime('%d.%m.%Y')}"
    
    file = BufferedInputFile(excel_content, filename=f"stats_transitions_{date_str}.xlsx")
    await message.answer_document(
        document=file,
        caption=f"Вот статистика за период: {period_text}",
        reply_markup=back_kb.as_markup()
    )
    await state.clear()


async def admin_stats_ratings(callback: types.CallbackQuery, state: FSMContext):
    """Рейтинги: период по умолчанию с начала текущего месяца по сегодня, сразу показываем выбор типа статистики."""
    await callback.answer()
    if not await check_admin_rights(callback.from_user.id):
        return
    now = datetime.now()
    period_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    period_end = now.replace(hour=23, minute=59, second=59, microsecond=999999)
    await state.update_data(period_start=period_start, period_end=period_end)
    await state.set_state(AdminPanelState.stats_period_type)
    text = "Выберите тип статистики для вывода"
    keyboard = _period_type_keyboard()
    try:
        await callback.message.edit_text(text, reply_markup=keyboard.as_markup())
    except Exception:
        await callback.message.answer(text, reply_markup=keyboard.as_markup())


async def admin_stats_period_ask_interval(callback: types.CallbackQuery, state: FSMContext):
    """Запрос интервала для статистики за период (при нажатии «Изменить период»)."""
    await callback.answer()
    if not await check_admin_rights(callback.from_user.id):
        return
    await state.set_state(AdminPanelState.stats_period_input)
    await state.update_data(period_ask_msg_id=callback.message.message_id)
    text = "Введите временной интервал для вывода статистики, в формате: 30.01.2026-20.02.2026"
    keyboard = InlineKeyboardBuilder()
    keyboard.row(InlineKeyboardButton(text="⬅️ Назад", callback_data="admin:stats"))
    try:
        await callback.message.edit_text(text, reply_markup=keyboard.as_markup())
    except Exception:
        await callback.message.answer(text, reply_markup=keyboard.as_markup())


def _period_type_keyboard():
    """Клавиатура выбора типа статистики за период."""
    keyboard = InlineKeyboardBuilder()
    keyboard.row(InlineKeyboardButton(text="Рейтинг по размещению", callback_data="admin:stats:period:top_placed"))
    keyboard.row(InlineKeyboardButton(text="Рейтинг по оценкам", callback_data="admin:stats:period:top_reviews"))
    keyboard.row(InlineKeyboardButton(text="⬅️ Назад", callback_data="admin:stats"))
    return keyboard


async def admin_stats_period_back(callback: types.CallbackQuery, state: FSMContext):
    """Возврат из экрана результата (рейтинги) к «Выберите тип статистики для вывода»."""
    await callback.answer()
    if not await check_admin_rights(callback.from_user.id):
        return
    data = await state.get_data()
    d_start, d_end = data.get("period_start"), data.get("period_end")
    if not d_start or not d_end:
        await state.clear()
        await admin_stats_menu(callback, state)
        return
    await state.set_state(AdminPanelState.stats_period_type)
    text = "Выберите тип статистики для вывода"
    keyboard = _period_type_keyboard()
    try:
        await callback.message.edit_text(text, reply_markup=keyboard.as_markup())
    except Exception:
        await callback.message.answer(text, reply_markup=keyboard.as_markup())
        try:
            await callback.message.delete()
        except Exception:
            pass


async def admin_stats_period_input(message: types.Message, state: FSMContext):
    """Обработка ввода интервала и показ выбора типа статистики за период (редактирование сообщения бота)."""
    d_start, d_end = _parse_period(message.text)
    back_kb = InlineKeyboardBuilder()
    back_kb.row(InlineKeyboardButton(text="⬅️ Назад", callback_data="admin:stats"))
    if not d_start or not d_end:
        err = "Неверный формат. Укажите интервал в формате ДД.ММ.ГГГГ-ДД.ММ.ГГГГ (например 30.01.2026-20.02.2026)."
        await _edit_or_send_period_step(message, state, err, back_kb)
        return
    if d_start > d_end:
        err = "Дата начала периода должна быть раньше даты окончания. Укажите интервал от меньшей даты к большей (например 01.01.2025-06.02.2026)."
        await _edit_or_send_period_step(message, state, err, back_kb)
        return
    try:
        await message.delete()
    except Exception:
        pass
    date_end = d_end.replace(hour=23, minute=59, second=59, microsecond=999999)
    await state.update_data(period_start=d_start, period_end=date_end)
    await state.set_state(AdminPanelState.stats_period_type)
    text = "Выберите тип статистики для вывода"
    keyboard = _period_type_keyboard()
    period_ask_msg_id = (await state.get_data()).get("period_ask_msg_id")
    if period_ask_msg_id:
        try:
            await bot.edit_message_text(
                chat_id=message.chat.id,
                message_id=period_ask_msg_id,
                text=text,
                reply_markup=keyboard.as_markup(),
            )
        except Exception:
            await message.answer(text, reply_markup=keyboard.as_markup())
    else:
        await message.answer(text, reply_markup=keyboard.as_markup())


async def _edit_or_send_period_step(message: types.Message, state: FSMContext, text: str, keyboard: InlineKeyboardBuilder):
    """Редактировать сообщение бота на шаге ввода периода или отправить новое; удалить сообщение пользователя."""
    try:
        await message.delete()
    except Exception:
        pass
    data = await state.get_data()
    period_ask_msg_id = data.get("period_ask_msg_id")
    kb = keyboard.as_markup() if keyboard else None
    if period_ask_msg_id:
        try:
            await bot.edit_message_text(chat_id=message.chat.id, message_id=period_ask_msg_id, text=text, reply_markup=kb)
        except Exception:
            await message.answer(text, reply_markup=kb)
    else:
        await message.answer(text, reply_markup=kb)


async def admin_stats_period_placed(callback: types.CallbackQuery, state: FSMContext):
    """Статистика «Размещено» за период."""
    await callback.answer()
    if not await check_admin_rights(callback.from_user.id):
        return
    data = await state.get_data()
    d_start, d_end = data.get("period_start"), data.get("period_end")
    if not d_start or not d_end:
        await state.clear()
        await admin_stats_menu(callback, state)
        return
    placed, sold, removed = await get_total_placed_sold_removed(d_start, d_end)
    top = await get_top_placed(d_start, d_end, 5)
    period_str = f"{d_start.strftime('%d.%m.%Y')}–{d_end.strftime('%d.%m.%Y')}"
    text = f"«Размещено»: {placed} за период {period_str}\n\nТОП-5 по размещению:\n"
    for tg_id, username, cnt in top:
        name_display = f"@{username}" if username else str(tg_id)
        text += f"{name_display} — {cnt}\n"
    keyboard = InlineKeyboardBuilder()
    keyboard.row(InlineKeyboardButton(text="📅 Изменить период", callback_data="admin:stats:period"))
    keyboard.row(InlineKeyboardButton(text="⬅️ Назад", callback_data="admin:stats:period:back"))
    try:
        await callback.message.edit_text(text, reply_markup=keyboard.as_markup())
    except Exception:
        await callback.message.answer(text, reply_markup=keyboard.as_markup())


async def admin_stats_period_sold(callback: types.CallbackQuery, state: FSMContext):
    """Статистика «Продано» за период."""
    await callback.answer()
    if not await check_admin_rights(callback.from_user.id):
        return
    data = await state.get_data()
    d_start, d_end = data.get("period_start"), data.get("period_end")
    if not d_start or not d_end:
        await state.clear()
        await admin_stats_menu(callback, state)
        return
    placed, sold, removed = await get_total_placed_sold_removed(d_start, d_end)
    top = await get_top_sold(d_start, d_end, 5)
    period_str = f"{d_start.strftime('%d.%m.%Y')}–{d_end.strftime('%d.%m.%Y')}"
    text = f"«Продано»: {sold} за период {period_str}\n\nТОП-5 по продажам:\n"
    for tg_id, username, cnt in top:
        name_display = f"@{username}" if username else str(tg_id)
        text += f"{name_display} — {cnt}\n"
    keyboard = InlineKeyboardBuilder()
    keyboard.row(InlineKeyboardButton(text="📅 Изменить период", callback_data="admin:stats:period"))
    keyboard.row(InlineKeyboardButton(text="⬅️ Назад", callback_data="admin:stats:period:back"))
    try:
        await callback.message.edit_text(text, reply_markup=keyboard.as_markup())
    except Exception:
        await callback.message.answer(text, reply_markup=keyboard.as_markup())


async def admin_stats_period_removed(callback: types.CallbackQuery, state: FSMContext):
    """Статистика «Снято» за период."""
    await callback.answer()
    if not await check_admin_rights(callback.from_user.id):
        return
    data = await state.get_data()
    d_start, d_end = data.get("period_start"), data.get("period_end")
    if not d_start or not d_end:
        await state.clear()
        await admin_stats_menu(callback, state)
        return
    placed, sold, removed = await get_total_placed_sold_removed(d_start, d_end)
    top = await get_top_removed(d_start, d_end, 5)
    period_str = f"{d_start.strftime('%d.%m.%Y')}–{d_end.strftime('%d.%m.%Y')}"
    text = f"«Снято»: {removed} за период {period_str}\n\nТОП-5 по снятию:\n"
    for tg_id, username, cnt in top:
        name_display = f"@{username}" if username else str(tg_id)
        text += f"{name_display} — {cnt}\n"
    keyboard = InlineKeyboardBuilder()
    keyboard.row(InlineKeyboardButton(text="📅 Изменить период", callback_data="admin:stats:period"))
    keyboard.row(InlineKeyboardButton(text="⬅️ Назад", callback_data="admin:stats:period:back"))
    try:
        await callback.message.edit_text(text, reply_markup=keyboard.as_markup())
    except Exception:
        await callback.message.answer(text, reply_markup=keyboard.as_markup())


async def admin_stats_period_top_placed(callback: types.CallbackQuery, state: FSMContext):
    """Рейтинг по размещению за период."""
    await callback.answer()
    if not await check_admin_rights(callback.from_user.id):
        return
    data = await state.get_data()
    d_start, d_end = data.get("period_start"), data.get("period_end")
    if not d_start or not d_end:
        await state.clear()
        await admin_stats_menu(callback, state)
        return
    placed, _, _ = await get_total_placed_sold_removed(d_start, d_end)
    top = await get_top_placed(d_start, d_end, 5)
    period_str = f"{d_start.strftime('%d.%m.%Y')}–{d_end.strftime('%d.%m.%Y')}"
    text = f"«Размещено» за период {period_str}: {placed}\n\nТОП-5:\n"
    for tg_id, username, cnt in top:
        name_display = f"@{username}" if username else str(tg_id)
        text += f"{name_display} — {cnt}\n"
    keyboard = InlineKeyboardBuilder()
    keyboard.row(InlineKeyboardButton(text="📅 Изменить период", callback_data="admin:stats:period"))
    keyboard.row(InlineKeyboardButton(text="⬅️ Назад", callback_data="admin:stats:period:back"))
    try:
        await callback.message.edit_text(text, reply_markup=keyboard.as_markup())
    except Exception:
        await callback.message.answer(text, reply_markup=keyboard.as_markup())


async def admin_stats_period_top_reviews(callback: types.CallbackQuery, state: FSMContext):
    """Рейтинг по оценкам активности за период."""
    await callback.answer()
    if not await check_admin_rights(callback.from_user.id):
        return
    data = await state.get_data()
    d_start, d_end = data.get("period_start"), data.get("period_end")
    if not d_start or not d_end:
        await state.clear()
        await admin_stats_menu(callback, state)
        return
    top = await get_top_reviews_activity(d_start, d_end, 5)
    period_str = f"{d_start.strftime('%d.%m.%Y')}–{d_end.strftime('%d.%m.%Y')}"
    text = f"«Оценки активности» за период {period_str}: отзывы покупателей продавцам.\n\nТОП-5 продавцов:\n"
    for tg_id, username, cnt, avg in top:
        name_display = f"@{username}" if username else str(tg_id)
        text += f"{name_display} — {cnt} отзывов (средний рейтинг {avg}/5)\n"
    keyboard = InlineKeyboardBuilder()
    keyboard.row(InlineKeyboardButton(text="📅 Изменить период", callback_data="admin:stats:period"))
    keyboard.row(InlineKeyboardButton(text="⬅️ Назад", callback_data="admin:stats:period:back"))
    try:
        await callback.message.edit_text(text, reply_markup=keyboard.as_markup())
    except Exception:
        await callback.message.answer(text, reply_markup=keyboard.as_markup())


